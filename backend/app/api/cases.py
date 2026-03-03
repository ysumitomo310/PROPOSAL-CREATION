"""案件管理API（TASK-D01）

POST /api/v1/cases — Excel取込 + 案件作成
GET  /api/v1/cases — 案件一覧
GET  /api/v1/cases/{case_id} — 案件詳細
"""

import json
import re
from io import BytesIO

import openpyxl
from fastapi import APIRouter, Depends, Form, HTTPException, UploadFile
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.dependencies import get_db
from app.models.case import Case
from app.models.requirement import FunctionalRequirement
from app.schemas.case import CaseResponse
from app.schemas.column_mapping import ColumnMapping, ColumnMappingConfig

router = APIRouter(prefix="/api/v1", tags=["cases"])

_MAX_FILE_SIZE = 50 * 1024 * 1024  # 50MB


@router.post("/cases", response_model=CaseResponse, status_code=201)
async def create_case(
    file: UploadFile,
    name: str = Form(...),
    product: str = Form(default="SAP"),
    column_mapping: str = Form(...),
    db: AsyncSession = Depends(get_db),
) -> CaseResponse:
    """案件作成 + Excel取込。"""
    # ファイル検証（拡張子ベース: ブラウザのMIMEタイプは信頼できないため）
    if file.filename and not file.filename.endswith(".xlsx"):
        raise HTTPException(422, f"Excel ファイル (.xlsx) のみ対応: {file.filename}")

    content = await file.read()
    if len(content) > _MAX_FILE_SIZE:
        raise HTTPException(413, "ファイルサイズが50MBを超えています")

    # column_mapping パース
    try:
        mapping_dict = json.loads(column_mapping)
        config = ColumnMappingConfig.model_validate(mapping_dict)
    except (json.JSONDecodeError, Exception) as e:
        raise HTTPException(422, f"column_mapping が不正です: {e}")

    # Excel パース
    try:
        wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
    except Exception as e:
        raise HTTPException(422, f"Excel ファイルを読み込めません: {e}")

    try:
        if config.sheet_name:
            if config.sheet_name not in wb.sheetnames:
                raise HTTPException(422, f"シート '{config.sheet_name}' が見つかりません")
            ws = wb[config.sheet_name]
        else:
            ws = wb.active

        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()

    if len(rows) < config.header_row:
        raise HTTPException(422, "ヘッダー行が見つかりません")

    # ヘッダーからカラムインデックスを構築
    header_row = rows[config.header_row - 1]
    headers = [str(h).strip() if h else "" for h in header_row]
    col_idx = _build_column_index(headers, config.columns)

    # データ行パース
    data_rows = rows[config.data_start_row - 1 :]
    requirement_rows = []
    for i, row in enumerate(data_rows):
        fn = _get_cell(row, col_idx.get("function_name"))
        if not fn:
            continue  # function_name が空の行はスキップ

        # business_category: 複数列を " > " で結合
        bc_parts = []
        for bc_col_name in (config.columns.business_category or []):
            bc_idx = _find_header_index(headers, bc_col_name)
            if bc_idx is not None:
                val = _get_cell(row, bc_idx)
                if val:
                    bc_parts.append(val)
        business_category = " > ".join(bc_parts) if bc_parts else None

        # importance: 値変換
        raw_importance = _get_cell(row, col_idx.get("importance"))
        if raw_importance and config.columns.importance_mapping:
            importance = config.columns.importance_mapping.get(
                str(raw_importance), str(raw_importance)
            )
        else:
            importance = raw_importance

        requirement_rows.append({
            "sequence_number": i + 1,
            "business_category": business_category,
            "business_name": _get_cell(row, col_idx.get("business_name")),
            "function_name": fn,
            "requirement_summary": _get_cell(row, col_idx.get("requirement_summary")),
            "requirement_detail": _get_cell(row, col_idx.get("requirement_detail")),
            "importance": importance,
            "original_row_json": {
                headers[j]: (str(v) if v is not None else None)
                for j, v in enumerate(row)
                if j < len(headers)
            },
        })

    if not requirement_rows:
        raise HTTPException(422, "有効な要件行が見つかりません")

    # 「同上」「〃」等の前行参照を解決
    requirement_rows = _resolve_forward_references(requirement_rows)

    # DB 保存
    case = Case(
        name=name,
        product=product,
        column_mapping=config.model_dump(),
        excel_filename=file.filename,
        total_requirements=len(requirement_rows),
    )
    db.add(case)
    await db.flush()

    for row_data in requirement_rows:
        fr = FunctionalRequirement(case_id=case.id, **row_data)
        db.add(fr)

    await db.commit()
    await db.refresh(case)

    return CaseResponse.model_validate(case)


@router.get("/cases", response_model=list[CaseResponse])
async def list_cases(
    db: AsyncSession = Depends(get_db),
) -> list[CaseResponse]:
    """案件一覧取得。"""
    stmt = select(Case).order_by(Case.created_at.desc())
    result = await db.execute(stmt)
    cases = result.scalars().all()
    return [CaseResponse.model_validate(c) for c in cases]


@router.get("/cases/{case_id}", response_model=CaseResponse)
async def get_case(
    case_id: str,
    db: AsyncSession = Depends(get_db),
) -> CaseResponse:
    """案件詳細取得。"""
    case = await db.get(Case, case_id)
    if not case:
        raise HTTPException(404, "案件が見つかりません")
    return CaseResponse.model_validate(case)


@router.delete("/cases/{case_id}", status_code=204)
async def delete_case(
    case_id: str,
    db: AsyncSession = Depends(get_db),
) -> None:
    """案件削除（要件・マッピング結果もカスケード削除）。"""
    case = await db.get(Case, case_id)
    if not case:
        raise HTTPException(404, "案件が見つかりません")
    await db.delete(case)
    await db.commit()


# ─── Internal helpers ───


def _find_header_index(headers: list[str], col_name: str | None) -> int | None:
    """ヘッダー名からインデックスを探す。"""
    if not col_name:
        return None
    for i, h in enumerate(headers):
        if h == col_name:
            return i
    return None


def _build_column_index(
    headers: list[str], columns: ColumnMapping
) -> dict[str, int | None]:
    """ColumnMapping から {field_name: header_index} を構築。"""
    return {
        "function_name": _find_header_index(headers, columns.function_name),
        "business_name": _find_header_index(headers, columns.business_name),
        "requirement_summary": _find_header_index(
            headers, columns.requirement_summary
        ),
        "requirement_detail": _find_header_index(
            headers, columns.requirement_detail
        ),
        "importance": _find_header_index(headers, columns.importance),
    }


def _get_cell(row: tuple, idx: int | None) -> str | None:
    """行からセル値を安全に取得。"""
    if idx is None or idx >= len(row):
        return None
    val = row[idx]
    if val is None:
        return None
    return str(val).strip() or None


# 「同上」「〃」等、前行と同じ値を示す表現にマッチする正規表現
_DITTO_PATTERN = re.compile(
    r"^(同上|同様|〃|同前|上記と同じ|上記に同じ|同じ|左記と同じ)$"
)

# 前行参照を解決するフィールド（function_name は key フィールドのため除外しない）
_RESOLVABLE_FIELDS = [
    "business_category",
    "business_name",
    "function_name",
    "requirement_summary",
    "requirement_detail",
    "importance",
]


def _resolve_forward_references(
    requirement_rows: list[dict],
) -> list[dict]:
    """「同上」「〃」「同様」等の前行参照をその内容で置換する。

    各フィールドごとに直近の非参照値を保持し、参照表現が出た場合に置換する。
    original_row_json は変更しない（元データ保持のため）。
    """
    prev: dict[str, str | None] = {f: None for f in _RESOLVABLE_FIELDS}

    for row in requirement_rows:
        for field in _RESOLVABLE_FIELDS:
            val = row.get(field)
            if val and _DITTO_PATTERN.match(val.strip()):
                # 前行参照 → 直近の有効値で置換
                row[field] = prev[field]
            elif val:
                # 有効な値 → 次行のために記憶
                prev[field] = val

    return requirement_rows
