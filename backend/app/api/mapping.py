"""マッピングAPI（TASK-D02）

POST /api/v1/cases/{case_id}/mapping/start — マッピング開始（202 Accepted）
GET  /api/v1/cases/{case_id}/mapping/stream — SSEストリーミング
GET  /api/v1/cases/{case_id}/mapping/results — 結果取得（フィルタ対応）
GET  /api/v1/cases/{case_id}/mapping/export — Excel一括エクスポート
"""

import json
import logging
from io import BytesIO
from urllib.parse import quote

from fastapi import APIRouter, BackgroundTasks, Depends, HTTPException, Query, Request
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import delete, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.dependencies import get_db
from app.models.case import Case
from app.models.mapping_result import MappingResult
from app.models.requirement import FunctionalRequirement
from app.schemas.mapping import (
    MappingResultDetail,
    MappingResultItem,
    MappingResultsResponse,
    MappingStartResponse,
)
from app.services.mapping.agent import MappingBatchProcessor

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/v1", tags=["mapping"])


def _get_or_build_graph(request: Request):
    """マッピンググラフを遅延初期化して返す。"""
    if not hasattr(request.app.state, "mapping_graph") or request.app.state.mapping_graph is None:
        from app.core.embedding import EmbeddingService
        from app.core.llm_client import LLMClient
        from app.core.neo4j_client import Neo4jClient
        from app.services.knowledge.search import HybridSearchService
        from app.services.mapping.agent import build_mapping_graph

        settings = request.app.state.settings
        llm_client = LLMClient(settings)
        neo4j_client = Neo4jClient(
            settings.NEO4J_URI, settings.NEO4J_USER, settings.NEO4J_PASSWORD
        )
        embedding_service = EmbeddingService(
            settings.OPENAI_API_KEY, settings.EMBEDDING_MODEL
        )
        search_service = HybridSearchService(neo4j_client, embedding_service)

        request.app.state.llm_client = llm_client
        request.app.state.neo4j_client = neo4j_client
        request.app.state.search_service = search_service
        request.app.state.mapping_graph = build_mapping_graph(
            llm_client, neo4j_client, search_service
        )

    return request.app.state.mapping_graph


@router.post("/cases/{case_id}/mapping/start", status_code=202)
async def start_mapping(
    case_id: str,
    request: Request,
    background_tasks: BackgroundTasks,
    db: AsyncSession = Depends(get_db),
) -> MappingStartResponse:
    """マッピング開始（非同期）→ 202 Accepted。"""
    case = await db.get(Case, case_id)
    if not case:
        raise HTTPException(404, "案件が見つかりません")
    if case.status == "mapping":
        raise HTTPException(409, "マッピングが既に実行中です")

    # 要件取得
    stmt = (
        select(FunctionalRequirement)
        .where(FunctionalRequirement.case_id == case_id)
        .order_by(FunctionalRequirement.sequence_number)
    )
    result = await db.execute(stmt)
    requirements = list(result.scalars().all())

    if not requirements:
        raise HTTPException(422, "マッピング対象の要件がありません")

    # 未完了 MappingResult を削除（再実行対応）
    # completed 以外（processing/error/pending）を削除してunique制約エラーを防ぐ
    req_ids = [r.id for r in requirements]
    await db.execute(
        delete(MappingResult).where(
            MappingResult.functional_requirement_id.in_(req_ids),
            MappingResult.status != "completed",
        )
    )

    # Case ステータス更新
    case.status = "mapping"
    await db.commit()

    # バッチプロセッサ作成
    graph = _get_or_build_graph(request)
    settings = request.app.state.settings
    processor = MappingBatchProcessor(
        graph=graph,
        session_factory=request.app.state.session_factory,
        max_concurrency=settings.MAPPING_MAX_CONCURRENCY,
        error_threshold=settings.MAPPING_ERROR_THRESHOLD,
    )
    request.app.state.active_batch_processors[case_id] = processor

    # バックグラウンドタスクで実行
    background_tasks.add_task(processor.run_batch, case_id, requirements)

    return MappingStartResponse(
        case_id=case_id,
        total_requirements=len(requirements),
    )


@router.get("/cases/{case_id}/mapping/stream")
async def stream_mapping(case_id: str, request: Request) -> StreamingResponse:
    """SSEストリーミングエンドポイント。"""
    processors: dict = getattr(request.app.state, "active_batch_processors", {})
    processor = processors.get(case_id)
    if not processor:
        raise HTTPException(404, "アクティブなマッピング処理が見つかりません")

    async def event_generator():
        async for event in processor.get_sse_events():
            yield event.format_sse()
            if event.type == "batch_complete":
                processors.pop(case_id, None)
                break

    return StreamingResponse(
        event_generator(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no",
        },
    )


@router.get("/cases/{case_id}/mapping/results", response_model=MappingResultsResponse)
async def get_mapping_results(
    case_id: str,
    judgment_level: str | None = Query(None),
    confidence: str | None = Query(None),
    importance: str | None = Query(None),
    status: str | None = Query(None),
    db: AsyncSession = Depends(get_db),
) -> MappingResultsResponse:
    """マッピング結果取得（フィルタ対応）。"""
    # Case 存在確認
    case = await db.get(Case, case_id)
    if not case:
        raise HTTPException(404, "案件が見つかりません")

    # JOIN クエリ構築
    stmt = (
        select(MappingResult, FunctionalRequirement)
        .join(
            FunctionalRequirement,
            MappingResult.functional_requirement_id == FunctionalRequirement.id,
        )
        .where(FunctionalRequirement.case_id == case_id)
    )

    # 動的フィルタ
    if judgment_level:
        stmt = stmt.where(MappingResult.judgment_level == judgment_level)
    if confidence:
        stmt = stmt.where(MappingResult.confidence == confidence)
    if importance:
        stmt = stmt.where(FunctionalRequirement.importance == importance)
    if status:
        stmt = stmt.where(MappingResult.status == status)

    stmt = stmt.order_by(FunctionalRequirement.sequence_number)

    result = await db.execute(stmt)
    rows = result.all()

    items = []
    completed_count = 0
    for mr, fr in rows:
        if mr.status == "completed":
            completed_count += 1
        items.append(
            MappingResultItem(
                id=str(mr.id),
                requirement_id=str(fr.id),
                sequence_number=fr.sequence_number,
                function_name=fr.function_name,
                requirement_summary=fr.requirement_summary,
                importance=fr.importance,
                judgment_level=mr.judgment_level,
                confidence=mr.confidence,
                confidence_score=mr.confidence_score,
                proposal_text=mr.proposal_text,
                rationale=mr.rationale,
                scope_item_analysis=mr.scope_item_analysis,
                gap_analysis=mr.gap_analysis,
                judgment_reason=mr.judgment_reason,
                matched_scope_items=mr.matched_scope_items,
                langsmith_trace_id=mr.langsmith_trace_id,
                status=mr.status,
            )
        )

    return MappingResultsResponse(
        case_id=case_id,
        total=case.total_requirements,
        completed=completed_count,
        results=items,
    )


@router.get("/cases/{case_id}/mapping/export")
async def export_mapping_results(
    case_id: str,
    db: AsyncSession = Depends(get_db),
) -> StreamingResponse:
    """マッピング結果をExcelファイルとしてエクスポート。

    1要件1行のフラット形式。
    """
    case = await db.get(Case, case_id)
    if not case:
        raise HTTPException(404, "案件が見つかりません")

    stmt = (
        select(MappingResult, FunctionalRequirement)
        .join(
            FunctionalRequirement,
            MappingResult.functional_requirement_id == FunctionalRequirement.id,
        )
        .where(FunctionalRequirement.case_id == case_id)
        .order_by(FunctionalRequirement.sequence_number)
    )
    result = await db.execute(stmt)
    rows = result.all()

    # ─── Excel生成 ───
    wb = Workbook()
    ws = wb.active
    ws.title = "マッピング結果"

    headers = [
        "No.", "業務分類", "業務名", "機能名", "要件概要", "要件詳細",
        "重要度", "判定レベル", "確信度", "確信度スコア",
        "提案文",
        "ScopeItem適合根拠", "ギャップ・課題", "判定結論",
        "マッチScope ID", "ステータス",
    ]

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    wrap_top    = Alignment(wrap_text=True, vertical="top")
    center_mid  = Alignment(horizontal="center", vertical="center")

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_mid

    for row_num, (mr, fr) in enumerate(rows, start=2):
        scope_ids = ", ".join(
            item.get("id", "") for item in (mr.matched_scope_items or [])
        )
        values = [
            fr.sequence_number,
            fr.business_category or "",
            fr.business_name or "",
            fr.function_name or "",
            fr.requirement_summary or "",
            fr.requirement_detail or "",
            fr.importance or "",
            mr.judgment_level or "",
            mr.confidence or "",
            mr.confidence_score,
            mr.proposal_text or "",
            mr.scope_item_analysis or mr.rationale or "",  # 旧データはratioanleをfallback
            mr.gap_analysis or "",
            mr.judgment_reason or "",
            scope_ids,
            mr.status or "",
        ]
        for col_idx, val in enumerate(values, 1):
            ws.cell(row=row_num, column=col_idx, value=val).alignment = wrap_top

    col_widths = [5, 10, 12, 24, 28, 35, 7, 16, 7, 9, 44, 44, 30, 20, 30, 9]
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

    last_col_letter = ws.cell(row=1, column=len(headers)).column_letter
    ws.auto_filter.ref = f"A1:{last_col_letter}1"
    ws.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"mapping_results_{case.name}.xlsx"
    encoded_filename = quote(filename, safe="")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}",
        },
    )
