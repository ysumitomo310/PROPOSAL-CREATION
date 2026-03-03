#!/usr/bin/env python3
"""PoC実行スクリプト（TASK-F03）

RFP Excelを読み込み、バックエンドAPIを経由せずに直接マッピングを実行する。
処理時間・エラー率・結果サマリーを出力。

Usage:
    python scripts/run_poc.py \
        --excel <path.xlsx> \
        --sheet "【要件一覧】" \
        --header-row 1 --data-start-row 3 \
        --function-name "小分類" \
        --requirement-summary "業務要件" \
        --requirement-detail "機能要件" \
        --importance "重要度" \
        --business-category "大分類" "中分類" \
        --importance-mapping "MUST=Must" "WANT=Should" \
        --local \
        [--limit N]
"""

import argparse
import asyncio
import json
import logging
import os
import sys
import time
from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker, create_async_engine

from app.core.config import get_settings
from app.core.embedding import EmbeddingService
from app.core.llm_client import LLMClient
from app.core.neo4j_client import Neo4jClient
from app.models.case import Case
from app.models.mapping_result import MappingResult
from app.models.requirement import FunctionalRequirement
from app.services.knowledge.search import HybridSearchService
from app.services.mapping.agent import MappingBatchProcessor, build_mapping_graph

logger = logging.getLogger(__name__)


def parse_excel(args) -> list[dict]:
    """RFP Excelを読み込み、要件行のリストを返す。"""
    wb = openpyxl.load_workbook(str(args.excel), data_only=True)

    if args.sheet:
        if args.sheet not in wb.sheetnames:
            raise ValueError(f"シート '{args.sheet}' が見つかりません。利用可能: {wb.sheetnames}")
        ws = wb[args.sheet]
    else:
        ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    # ヘッダー解析
    header_row = rows[args.header_row - 1]
    headers = [str(h).strip() if h else "" for h in header_row]

    def find_col(name):
        for i, h in enumerate(headers):
            if h == name:
                return i
        return None

    fn_idx = find_col(args.function_name)
    rs_idx = find_col(args.requirement_summary) if args.requirement_summary else None
    rd_idx = find_col(args.requirement_detail) if args.requirement_detail else None
    imp_idx = find_col(args.importance) if args.importance else None
    bc_indices = [find_col(bc) for bc in (args.business_category or [])]
    bc_indices = [i for i in bc_indices if i is not None]

    if fn_idx is None:
        raise ValueError(f"機能名列 '{args.function_name}' が見つかりません。ヘッダー: {headers}")

    # importance_mapping パース
    imp_map = {}
    if args.importance_mapping:
        for mapping in args.importance_mapping:
            k, v = mapping.split("=", 1)
            imp_map[k.strip()] = v.strip()

    # データ行パース
    data_rows = rows[args.data_start_row - 1:]
    requirements = []
    for i, row in enumerate(data_rows):
        fn = row[fn_idx] if fn_idx < len(row) else None
        if not fn:
            continue

        fn = str(fn).strip()
        rs = str(row[rs_idx]).strip() if rs_idx and rs_idx < len(row) and row[rs_idx] else None
        rd = str(row[rd_idx]).strip() if rd_idx and rd_idx < len(row) and row[rd_idx] else None

        # business_category
        bc_parts = []
        for bc_i in bc_indices:
            if bc_i < len(row) and row[bc_i]:
                bc_parts.append(str(row[bc_i]).strip())
        bc = " > ".join(bc_parts) if bc_parts else None

        # importance
        raw_imp = str(row[imp_idx]).strip() if imp_idx and imp_idx < len(row) and row[imp_idx] else None
        importance = imp_map.get(raw_imp, raw_imp) if raw_imp else None

        requirements.append({
            "sequence_number": i + 1,
            "function_name": fn,
            "requirement_summary": rs,
            "requirement_detail": rd,
            "business_category": bc,
            "importance": importance,
        })

    if args.limit and args.limit > 0:
        requirements = requirements[:args.limit]

    return requirements


async def run_poc(args):
    """PoC マッピング実行。"""
    # ローカル接続設定
    if args.local:
        os.environ["POSTGRES_HOST"] = "localhost"
        os.environ["POSTGRES_PORT"] = "5435"
        os.environ["NEO4J_URI"] = "bolt://localhost:7688"

    settings = get_settings()

    # LangSmithを無効化（PoC実行ではキー未設定の場合がある）
    if not settings.langsmith_enabled:
        os.environ["LANGCHAIN_TRACING_V2"] = "false"

    # LangChain init_chat_model() は環境変数からAPIキーを読むためセット
    if settings.OPENAI_API_KEY:
        os.environ["OPENAI_API_KEY"] = settings.OPENAI_API_KEY
    if settings.ANTHROPIC_API_KEY:
        os.environ["ANTHROPIC_API_KEY"] = settings.ANTHROPIC_API_KEY

    # Excel パース
    print("=== PoC Mapping Execution ===")
    print(f"Excel: {args.excel}")
    print(f"Sheet: {args.sheet or '(active)'}")
    requirements_data = parse_excel(args)
    print(f"要件数: {len(requirements_data)}")
    print()

    if not requirements_data:
        print("要件が0件です。終了。")
        return

    # DB接続
    engine = create_async_engine(settings.database_url, echo=False)
    session_factory = async_sessionmaker(engine, class_=AsyncSession, expire_on_commit=False)

    # Case + FunctionalRequirements 作成
    async with session_factory() as session:
        case = Case(
            name=f"PoC_{Path(args.excel).stem}_{datetime.now().strftime('%H%M%S')}",
            product="SAP",
            status="mapping",
            total_requirements=len(requirements_data),
        )
        session.add(case)
        await session.flush()
        case_id = case.id

        fr_objects = []
        for rd in requirements_data:
            fr = FunctionalRequirement(case_id=case_id, **rd)
            session.add(fr)
            fr_objects.append(fr)

        await session.commit()

        # refresh してIDを取得
        for fr in fr_objects:
            await session.refresh(fr)

    print(f"Case作成: {case_id}")
    print(f"FunctionalRequirements: {len(fr_objects)}件")

    # マッピンググラフ構築
    print("\nサービス初期化中...")
    llm_client = LLMClient(settings)
    neo4j_client = Neo4jClient(
        settings.NEO4J_URI, settings.NEO4J_USER, settings.NEO4J_PASSWORD
    )
    embedding_service = EmbeddingService(
        settings.OPENAI_API_KEY, settings.EMBEDDING_MODEL
    )
    search_service = HybridSearchService(neo4j_client, embedding_service)
    graph = build_mapping_graph(llm_client, neo4j_client, search_service)

    # バッチ実行
    print(f"マッピング開始（並行数: {settings.MAPPING_MAX_CONCURRENCY}）...")
    start_time = time.time()

    processor = MappingBatchProcessor(
        graph=graph,
        session_factory=session_factory,
        max_concurrency=settings.MAPPING_MAX_CONCURRENCY,
        error_threshold=settings.MAPPING_ERROR_THRESHOLD,
    )

    # SSEイベント収集（進捗表示用）
    events = []

    async def show_progress():
        async for event in processor.get_sse_events():
            events.append(event)
            if event.type == "requirement_complete":
                d = event.data
                print(
                    f"  [{d.get('completed_count', '?')}/{len(fr_objects)}] "
                    f"{d.get('function_name', '')[:30]} → {d.get('judgment_level', '?')} "
                    f"({d.get('confidence', '?')})"
                )
            elif event.type == "error":
                print(f"  ERROR: {event.data.get('error', '')[:80]}")
            elif event.type == "batch_complete":
                break

    # 並行実行
    result = await asyncio.gather(
        processor.run_batch(case_id, fr_objects),
        show_progress(),
    )
    batch_result = result[0]

    elapsed = time.time() - start_time

    # 結果サマリー
    print(f"\n{'=' * 60}")
    print("=== PoC Result Summary ===")
    print(f"{'=' * 60}")
    print(f"Case ID:          {case_id}")
    print(f"総要件数:          {batch_result.total}")
    print(f"完了:              {batch_result.completed}")
    print(f"エラー:            {batch_result.errors}")
    print(f"中止:              {'Yes' if batch_result.aborted else 'No'}")
    print(f"処理時間:          {elapsed:.1f}秒")
    print(f"1件あたり平均:     {elapsed / batch_result.total:.1f}秒")
    print(f"エラー率:          {batch_result.errors / batch_result.total * 100:.1f}%")

    # 判定レベル分布
    req_events = [e for e in events if e.type == "requirement_complete"]
    judgment_counts = {}
    confidence_counts = {}
    for e in req_events:
        jl = e.data.get("judgment_level", "Unknown")
        conf = e.data.get("confidence", "Unknown")
        judgment_counts[jl] = judgment_counts.get(jl, 0) + 1
        confidence_counts[conf] = confidence_counts.get(conf, 0) + 1

    print(f"\n判定レベル分布:")
    for jl, cnt in sorted(judgment_counts.items()):
        print(f"  {jl}: {cnt}件")

    print(f"\n確信度分布:")
    for conf, cnt in sorted(confidence_counts.items()):
        print(f"  {conf}: {cnt}件")

    print(f"{'=' * 60}")

    # JSON出力
    output = {
        "case_id": case_id,
        "executed_at": datetime.now(timezone.utc).isoformat(),
        "excel_file": str(args.excel),
        "total_requirements": batch_result.total,
        "completed": batch_result.completed,
        "errors": batch_result.errors,
        "aborted": batch_result.aborted,
        "elapsed_seconds": round(elapsed, 1),
        "avg_seconds_per_requirement": round(elapsed / batch_result.total, 1),
        "error_rate": round(batch_result.errors / batch_result.total * 100, 1),
        "judgment_distribution": judgment_counts,
        "confidence_distribution": confidence_counts,
    }

    output_path = Path(args.excel).parent / f"poc_result_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
    output_path.write_text(json.dumps(output, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"\n結果JSON: {output_path}")

    # クリーンアップ
    await neo4j_client.close()
    await engine.dispose()


def main():
    parser = argparse.ArgumentParser(description="PoC Mapping Execution")
    parser.add_argument("--excel", type=Path, required=True, help="RFP Excel パス")
    parser.add_argument("--sheet", type=str, default=None, help="シート名")
    parser.add_argument("--header-row", type=int, default=1, help="ヘッダー行番号")
    parser.add_argument("--data-start-row", type=int, default=2, help="データ開始行番号")
    parser.add_argument("--function-name", type=str, required=True, help="機能名列ヘッダー")
    parser.add_argument("--requirement-summary", type=str, default=None, help="要件概要列ヘッダー")
    parser.add_argument("--requirement-detail", type=str, default=None, help="要件詳細列ヘッダー")
    parser.add_argument("--importance", type=str, default=None, help="重要度列ヘッダー")
    parser.add_argument("--business-category", nargs="*", default=None, help="業務分類列ヘッダー（複数可）")
    parser.add_argument("--importance-mapping", nargs="*", default=None, help="重要度マッピング（例: MUST=Must WANT=Should）")
    parser.add_argument("--limit", type=int, default=0, help="処理件数上限（0=全件）")
    parser.add_argument("--local", action="store_true", help="ローカル実行（localhost接続）")

    args = parser.parse_args()

    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    )
    # サードパーティログを抑制
    logging.getLogger("httpx").setLevel(logging.WARNING)
    logging.getLogger("httpcore").setLevel(logging.WARNING)
    logging.getLogger("openai").setLevel(logging.WARNING)

    asyncio.run(run_poc(args))


if __name__ == "__main__":
    main()
